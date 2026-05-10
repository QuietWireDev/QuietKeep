# QuietKeep: routers/patches.py
# Scan, patch, reboot, and history endpoints.
# Author: QuietWire (Dennis Ayotte)

import csv
import io
import re
import logging
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import (
    HostModel,
    PatchHistoryModel,
    PatchHistoryResponse,
    PatchRequest,
)
from app.services.activity import log_activity
from app.services.patcher import patch_host
from app.services.scanner import scan_all_hosts, scan_host
from app.ssh.client import ssh_client

router = APIRouter(prefix="/api", tags=["patches"])
logger = logging.getLogger(__name__)


@router.post("/scan")
async def trigger_scan_all(db: AsyncSession = Depends(get_db)):
    """Trigger a scan of all hosts."""
    results = await scan_all_hosts(db)
    online = sum(1 for r in results if r.get('status') != 'error')
    await log_activity(db, 'scan', f'Fleet scan complete: {online}/{len(results)} hosts responded')
    return {"message": "Scan complete", "results": results}


@router.post("/scan/{host_id}")
async def trigger_scan_host(host_id: int, db: AsyncSession = Depends(get_db)):
    """Trigger a scan of a single host."""
    result = await db.execute(select(HostModel).where(HostModel.id == host_id))
    host = result.scalar_one_or_none()
    if not host:
        raise HTTPException(status_code=404, detail="Host not found")
    scan_result = await scan_host(host, db)
    await log_activity(db, 'scan', f'Scanned {host.hostname}', host_id=host.id, hostname=host.hostname)
    return {"message": f"Scan complete for {host.hostname}", "result": scan_result}


@router.post("/patch")
async def trigger_patch(request: PatchRequest, db: AsyncSession = Depends(get_db)):
    """Patch selected hosts."""
    results = []
    for host_id in request.host_ids:
        result = await db.execute(select(HostModel).where(HostModel.id == host_id))
        host = result.scalar_one_or_none()
        if not host:
            results.append({"host_id": host_id, "status": "not_found"})
            continue
        history = await patch_host(host, db)
        results.append({
            "host_id": host.id,
            "hostname": host.hostname,
            "status": history.status,
            "packages_updated": history.packages_updated,
        })
    return {"message": "Patching complete", "results": results}


@router.post("/patch/all")
async def trigger_patch_all(db: AsyncSession = Depends(get_db)):
    """Patch all online, patch-eligible hosts (standard upgrade only, no kernel).

    Returns per-host results so the frontend can surface successes and failures.
    """
    result = await db.execute(
        select(HostModel).where(
            HostModel.is_online == True,
            HostModel.is_patch_target == True,
            HostModel.pending_updates > 0,
        )
    )
    hosts = result.scalars().all()
    if not hosts:
        return {"message": "No hosts with pending updates", "results": []}

    results = []
    for host in hosts:
        try:
            history = await patch_host(host, db)
            results.append({
                "host_id": host.id,
                "hostname": host.hostname,
                "status": history.status,
                "packages_updated": history.packages_updated,
            })
        except Exception as e:
            logger.exception("Bulk patch failed for host_id=%s hostname=%s", host.id, host.hostname)
            results.append({
                "host_id": host.id,
                "hostname": host.hostname,
                "status": "error",
                "packages_updated": 0,
                "error": "Patching failed due to an internal error",
            })

    # Re-scan all patched hosts so pending counts refresh immediately
    for host in hosts:
        try:
            await scan_host(host, db)
        except Exception:
            pass  # scan failure shouldn't block results

    succeeded = sum(1 for r in results if r['status'] == 'success')
    failed = sum(1 for r in results if r['status'] in ('failed', 'error'))
    total_pkgs = sum(r['packages_updated'] for r in results)
    await log_activity(db, 'patch', f'Bulk patch: {succeeded} succeeded, {failed} failed, {total_pkgs} pkgs updated')
    return {"message": "Bulk patching complete", "results": results}


@router.post("/patch/{host_id}")
async def trigger_patch_host(host_id: int, db: AsyncSession = Depends(get_db)):
    """Patch a single host."""
    result = await db.execute(select(HostModel).where(HostModel.id == host_id))
    host = result.scalar_one_or_none()
    if not host:
        raise HTTPException(status_code=404, detail="Host not found")
    history = await patch_host(host, db)
    await log_activity(db, 'patch', f'Patched {host.hostname}: {history.status}, {history.packages_updated} pkgs', host_id=host.id, hostname=host.hostname)
    return {
        "message": f"Patching complete for {host.hostname}",
        "status": history.status,
        "packages_updated": history.packages_updated,
    }


@router.post("/patch/{host_id}/install-held-back")
async def trigger_install_held_back(host_id: int, db: AsyncSession = Depends(get_db)):
    """Install the packages the last `apt-get upgrade` kept back.

    Runs `apt-get upgrade --with-new-pkgs` via the same patcher path as a
    regular patch, which pulls in new versioned subpackages (typically a
    new kernel image) without performing any removals. Creates a normal
    patch-history entry so the user sees this run alongside their other
    patches, and clears host.held_back_packages on success so the UI no
    longer shows the follow-up card.

    The operation usually leaves the host wanting a reboot; QuietKeep's
    existing reboot detection picks that up on the next scan and lights
    up the Reboot button naturally.
    """
    result = await db.execute(select(HostModel).where(HostModel.id == host_id))
    host = result.scalar_one_or_none()
    if not host:
        raise HTTPException(status_code=404, detail="Host not found")
    if host.os_type not in ("apt", "proxmox"):
        # kali already uses dist-upgrade and pacman does full upgrades, so
        # these OS types never end up with held-back packages to install.
        raise HTTPException(
            status_code=400,
            detail=f"Install-held-back does not apply to os_type={host.os_type}",
        )
    history = await patch_host(host, db, include_new_pkgs=True)
    return {
        "message": f"Held-back install complete for {host.hostname}",
        "status": history.status,
        "packages_updated": history.packages_updated,
    }


@router.post("/reboot/{host_id}")
async def trigger_reboot(host_id: int, db: AsyncSession = Depends(get_db)):
    """Reboot a single host."""
    result = await db.execute(select(HostModel).where(HostModel.id == host_id))
    host = result.scalar_one_or_none()
    if not host:
        raise HTTPException(status_code=404, detail="Host not found")
    if not host.is_patch_target:
        raise HTTPException(status_code=400, detail=f"{host.hostname} is not a patch target")
    if not host.reboot_required:
        raise HTTPException(status_code=400, detail=f"{host.hostname} does not need a reboot")

    success, message = await ssh_client.reboot_host(host.ip_address, host.username)
    if success:
        # Mark offline immediately. Host won't respond until it comes back up.
        # reboot_required clears now; next scan will re-detect if still needed.
        host.reboot_required = False
        host.is_online = False
        await db.commit()
        await log_activity(db, 'reboot', f'Rebooted {host.hostname}', host_id=host.id, hostname=host.hostname)
    return {"message": message, "hostname": host.hostname, "success": success}


@router.get("/history", response_model=list[PatchHistoryResponse])
async def get_history(limit: int = 50, db: AsyncSession = Depends(get_db)):
    """Get recent patch history across all hosts."""
    result = await db.execute(
        select(PatchHistoryModel)
        .order_by(PatchHistoryModel.started_at.desc())
        .limit(limit)
    )
    return result.scalars().all()


@router.get("/history/{host_id}", response_model=list[PatchHistoryResponse])
async def get_host_history(host_id: int, db: AsyncSession = Depends(get_db)):
    """Get patch history for a specific host."""
    result = await db.execute(
        select(PatchHistoryModel)
        .where(PatchHistoryModel.host_id == host_id)
        .order_by(PatchHistoryModel.started_at.desc())
    )
    return result.scalars().all()


_HEADER_COLS = ["started_at", "completed_at", "status", "packages_updated", "packages"]


def _parse_package_names(log_output: str | None) -> str:
    """Extract upgraded package names from apt/pacman log output.

    For apt: parses "Unpacking <name> (<version>)" lines.
    Returns a semicolon-separated string of package names, or empty string.
    """
    if not log_output:
        return ""
    names: list[str] = []
    for line in log_output.splitlines():
        stripped = line.strip()
        # apt: "Unpacking libfoo:amd64 (1.2.3-1) over ..."
        if stripped.startswith("Unpacking "):
            parts = stripped.split()
            if len(parts) >= 2:
                # Remove architecture suffix like ":amd64"
                pkg = parts[1].split(":")[0]
                if pkg not in names:
                    names.append(pkg)
        # pacman: "upgrading foo"
        elif stripped.startswith("upgrading "):
            pkg = stripped.split()[1].rstrip(".")
            if pkg not in names:
                names.append(pkg)
    return "; ".join(names)


def _history_to_csv(rows: list, hosts_map: dict[int, str]) -> str:
    """Convert patch history rows to CSV string."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["hostname"] + _HEADER_COLS)
    for r in rows:
        writer.writerow([
            hosts_map.get(r.host_id, "unknown"),
            r.started_at.isoformat() if r.started_at else "",
            r.completed_at.isoformat() if r.completed_at else "",
            r.status,
            r.packages_updated,
            _parse_package_names(r.log_output),
        ])
    return buf.getvalue()


def _build_xlsx(rows_by_host: dict[str, list]) -> io.BytesIO:
    """Build an xlsx workbook with one sheet per host.

    Each patch run is expanded to one row per package, with the patch
    metadata (dates, status, count) repeated on every row for easy
    filtering and pivoting. Runs with 0 packages still get one row.
    """
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")

    for hostname in sorted(rows_by_host.keys()):
        rows = rows_by_host[hostname]
        # Sheet names max 31 chars in xlsx
        sheet_name = hostname[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write header
        for col_idx, col_name in enumerate(_HEADER_COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        # Write rows — one row per package
        row_idx = 2
        for r in rows:
            started = r.started_at.isoformat() if r.started_at else ""
            completed = r.completed_at.isoformat() if r.completed_at else ""
            pkg_str = _parse_package_names(r.log_output)
            packages = [p.strip() for p in pkg_str.split(";") if p.strip()] if pkg_str else []

            if not packages:
                # No packages — still write one row for the run
                ws.cell(row=row_idx, column=1, value=started)
                ws.cell(row=row_idx, column=2, value=completed)
                ws.cell(row=row_idx, column=3, value=r.status)
                ws.cell(row=row_idx, column=4, value=r.packages_updated)
                ws.cell(row=row_idx, column=5, value="")
                row_idx += 1
            else:
                for pkg in packages:
                    ws.cell(row=row_idx, column=1, value=started)
                    ws.cell(row=row_idx, column=2, value=completed)
                    ws.cell(row=row_idx, column=3, value=r.status)
                    ws.cell(row=row_idx, column=4, value=r.packages_updated)
                    ws.cell(row=row_idx, column=5, value=pkg)
                    row_idx += 1

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/history/export/xlsx")
async def export_history_xlsx(db: AsyncSession = Depends(get_db)):
    """Export all patch history as an Excel file with one sheet per host."""
    result = await db.execute(
        select(PatchHistoryModel).order_by(PatchHistoryModel.started_at.desc())
    )
    rows = result.scalars().all()
    host_ids = list({r.host_id for r in rows})
    hosts_result = await db.execute(select(HostModel).where(HostModel.id.in_(host_ids)))
    hosts_map = {h.id: h.hostname for h in hosts_result.scalars().all()}

    rows_by_host: dict[str, list] = defaultdict(list)
    for r in rows:
        hostname = hosts_map.get(r.host_id, "unknown")
        rows_by_host[hostname].append(r)

    xlsx_buf = _build_xlsx(rows_by_host)
    return StreamingResponse(
        xlsx_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quietkeep-patch-history.xlsx"},
    )


@router.get("/history/{host_id}/export/csv")
async def export_host_history_csv(host_id: int, db: AsyncSession = Depends(get_db)):
    """Export patch history for a single host as a CSV file."""
    host_result = await db.execute(select(HostModel).where(HostModel.id == host_id))
    host = host_result.scalar_one_or_none()
    if not host:
        raise HTTPException(status_code=404, detail="Host not found")
    result = await db.execute(
        select(PatchHistoryModel)
        .where(PatchHistoryModel.host_id == host_id)
        .order_by(PatchHistoryModel.started_at.desc())
    )
    rows = result.scalars().all()
    csv_content = _history_to_csv(rows, {host_id: host.hostname})
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=quietkeep-patch-history-{host.hostname}.csv"},
    )
